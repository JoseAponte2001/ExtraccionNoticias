from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from subprocess import CREATE_NO_WINDOW
from openpyxl import Workbook
global driver
import time
import os

import sys
print("Python Version: " + str(sys.version_info))
service = Service()
service.creationflags = CREATE_NO_WINDOW
opts = Options()
opts.binary_location = 'C:\\Program Files\\Mozilla Firefox\\firefox.exe'
opts.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/csv")

def eltiempo():
    with open('./links_el_tiempo.txt', 'w') as f:
        f.write("")
    driver.get("https://www.eltiempo.com/noticias/nevado-del-ruiz")
    elements = driver.find_elements("xpath",'//div[@class="col1"]//a[(@class="title page-link")]')
    for element in elements:
        href = element.get_attribute("href")
        with open("./links_el_tiempo.txt", "a") as f:
            f.write(href+'\n')
    archivo = open('./links_el_tiempo.txt', 'r')
    wb = Workbook()
    hoja_activa = wb.active
    fila = 2
    for link in archivo:
        try:
            driver.get(link)
            celda_titulo = 'B'+str(fila)
            titulo_elemento = driver.find_element("xpath", '//div[@class="titulo-principal-bk"]')
            titulo = titulo_elemento.text
            hoja_activa[celda_titulo] = titulo

            celda_contenido = 'C'+str(fila)
            texto = ''
            texto_elementos = driver.find_elements("xpath", '//p[@class="contenido"]')
            for element in texto_elementos:
                texto = texto + element.text
            hoja_activa[celda_contenido] = texto
            if texto == "¿Te gusta estar informado? Disfruta del mejor contenido sin límites. Suscríbete aquí.":
                texto = ''
                texto_elementos = driver.find_elements("xpath", '//p[@class="contenido "]')
                for element in texto_elementos:
                    texto = texto + element.text
                hoja_activa[celda_contenido] = texto
            """
            celda_fecha = 'D' + str(fila)
            fecha_elemento = driver.find_elements("xpath", '//*[@class="publishedAt"]')
            fecha = fecha_elemento[1].text
            for elemento in fecha_elemento:
                if elemento.text != '':
                    fecha = elemento.text
            hoja_activa[celda_fecha] = fecha
            """
            celda_link = 'A'+str(fila)
            hoja_activa[celda_link] = link
            fila = fila + 1
        except:
            print("se rompio")
    wb.save('eltiempo.xlsx')

def semana():
    with open('./links_semana.txt', 'w') as f:
        f.write("")
    driver.get("https://www.semana.com/noticias/volcan-nevado-del-ruiz/")
    ver_mas = driver.find_element("xpath", "//h1[contains(text(), 'Ver más')]")
    for i in range(1, 16):
        ver_mas.location_once_scrolled_into_view
        driver.execute_script("window.scrollBy(0, -60)")
        ver_mas.click()
    elements = driver.find_elements("xpath",'//a[h2[@class="card-title h3"]]')
    for element in elements:
        href = element.get_attribute("href")
        with open("./links_semana.txt", "a") as f:
            f.write(href+'\n')
    archivo = open('./links_semana.txt', 'r')
    wb = Workbook()
    hoja_activa = wb.active
    fila = 2
    for link in archivo:
        try:
            driver.get(link)
            celda_titulo = 'B'+str(fila)
            titulo_elemento = driver.find_element("xpath", '//div[@class="article-header-box"]//h1//span')
            titulo = titulo_elemento.text
            hoja_activa[celda_titulo] = titulo

            celda_contenido = 'C'+str(fila)
            texto = ''

            texto_elemento = driver.find_element("xpath", '//article')
            texto = texto_elemento.text
            hoja_activa[celda_contenido] = texto

            """
            celda_fecha = 'B' + str(fila)
            fecha_elemento = driver.find_element("xpath", '//div[@class="datetime"]')
            fecha = fecha_elemento.text
            hoja_activa[celda_fecha] = fecha
            """

            celda_link = 'A'+str(fila)
            hoja_activa[celda_link] = link
            fila = fila + 1
        except Exception as e:
            print(f"Se produjo una excepción: {e}")
    wb.save('semana.xlsx')

def elespectador():
    with open('./links_el_espectador.txt', 'w') as f:
        f.write("")
    driver.get("https://www.elespectador.com/tags/nevado-del-ruiz/")
    ver_mas = driver.find_element("xpath", "//div[contains(text(), 'Cargar más notas')]")
    for i in range(1, 10):
        ver_mas.location_once_scrolled_into_view
        driver.execute_script("window.scrollBy(0, -60)")
        ver_mas.click()
    elements = driver.find_elements("xpath", '//div[@class="Card-Container"]//h2//a')
    for element in elements:
        href = element.get_attribute("href")
        with open("./links_el_espectador.txt", "a") as f:
            f.write(href+'\n')
    archivo = open('./links_el_espectador.txt', 'r')
    wb = Workbook()
    hoja_activa = wb.active
    fila = 2
    for link in archivo:
        try:
            driver.get(link)
            celda_titulo = 'B'+str(fila)
            titulo_elemento = driver.find_element("xpath", '//h1[@class="Title ArticleHeader-Title Title_article"]')
            titulo = titulo_elemento.text
            hoja_activa[celda_titulo] = titulo

            celda_contenido = 'C'+str(fila)
            texto = ''
            texto_elemento = driver.find_element("xpath", '//section')
            texto = texto_elemento.text
            hoja_activa[celda_contenido] = texto

            celda_link = 'A'+str(fila)
            hoja_activa[celda_link] = link
            fila = fila + 1
        except Exception as e:
            print(f"Se produjo una excepción: {e}")
    wb.save('el_espectador.xlsx')
try:
    driver = webdriver.Firefox(service=service, options=opts)

    extension_path = os.path.abspath("extension/adblock.xpi")
    driver.install_addon(extension_path, temporary=True)

    time.sleep(2)

    driver.switch_to.window(driver.window_handles[0])


    elespectador()
    driver.quit()
except Exception:
    e = sys.exc_info()[1]
    print(e)

